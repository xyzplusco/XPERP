#!/usr/bin/env python3
"""Audit XP ERP source workbooks and write compact review reports.

The goal is not to perfectly import every row yet. The goal is to keep the
product honest by making source shape, missing fields, and operational signals
visible before schema or UI work drifts into generic CRM territory.
"""

from __future__ import annotations

import collections
import datetime as dt
import hashlib
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW = DATA / "raw"
DOCS = ROOT / "docs"

SOURCES = {
    "cleaned_partners": DATA / "XP_partner_list_cleaned_DB_ready.xlsx",
    "deal_list": DATA / "XP Deal list_대외비_20260805.xlsx",
    "network_original": RAW / "00.XP_파트너 및 네트워크 리스트_260813.xlsx",
    "to_go": RAW / "To Go List XYZ Plus (7).xlsx",
}

XML_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return str(value).strip()


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def counter_dict(counter: collections.Counter[str], limit: int | None = None) -> dict[str, int]:
    items = counter.most_common(limit)
    return {key: count for key, count in items}


def audit_cleaned_partners(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Cleaned_Partners"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [cell_text(v) for v in rows[0]]
    idx = {name: i for i, name in enumerate(headers) if name}
    data_rows = rows[1:]

    def col(row: tuple[Any, ...], name: str) -> str:
        pos = idx.get(name)
        return cell_text(row[pos]) if pos is not None and pos < len(row) else ""

    output = {
        "file": str(path.relative_to(ROOT)),
        "sha256": file_sha256(path),
        "sheet": "Cleaned_Partners",
        "rows": len(data_rows),
        "columns": len(headers),
        "category_counts": counter_dict(collections.Counter(col(r, "category") or "(blank)" for r in data_rows)),
        "nda_counts": counter_dict(collections.Counter(col(r, "nda_status") or "(blank)" for r in data_rows)),
        "agreement_counts": counter_dict(collections.Counter(col(r, "agreement_status") or "(blank)" for r in data_rows)),
        "has_personal_email": sum(1 for r in data_rows if col(r, "email_personal")),
        "has_company": sum(1 for r in data_rows if col(r, "company")),
        "has_personal_phone": sum(1 for r in data_rows if col(r, "phone_personal")),
        "sample_names": [col(r, "name") for r in data_rows[:10]],
    }
    wb.close()
    return output


def header_map(ws: Any, header_row: int) -> dict[str, int]:
    headers = [cell_text(c.value) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    return {name: i for i, name in enumerate(headers) if name}


def audit_deal_sheet(path: Path, sheet_name: str, header_row: int) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    idx = header_map(ws, header_row)
    headers = list(idx.keys())

    def col(row: tuple[Any, ...], name: str) -> str:
        pos = idx.get(name)
        return cell_text(row[pos]) if pos is not None and pos < len(row) else ""

    data_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(cell_text(v) for v in row):
            data_rows.append(row)

    company_rows = [r for r in data_rows if col(r, "회사명")]
    update_columns = [h for h in headers if re.search(r"\d+월|주", h)]
    output = {
        "sheet": sheet_name,
        "header_row": header_row,
        "non_empty_rows": len(data_rows),
        "rows_with_company": len(company_rows),
        "unique_companies": len({col(r, "회사명") for r in company_rows}),
        "service_sector_counts": counter_dict(collections.Counter(col(r, "서비스 섹터") or "(blank)" for r in company_rows), 20),
        "pl_nonblank": sum(1 for r in company_rows if col(r, "PL")),
        "pm1_nonblank": sum(1 for r in company_rows if col(r, "PM 1")),
        "pm2_nonblank": sum(1 for r in company_rows if col(r, "PM2")),
        "weekly_update_columns": len(update_columns),
        "sample_companies": [col(r, "회사명") for r in company_rows[:15]],
    }
    wb.close()
    return output


def audit_deal_list(path: Path) -> dict[str, Any]:
    return {
        "file": str(path.relative_to(ROOT)),
        "sha256": file_sha256(path),
        "sheets": [
            audit_deal_sheet(path, "Deals_0731", 7),
            audit_deal_sheet(path, "M&A Deal_0809", 3),
            audit_deal_sheet(path, "Deals_투자매각", 5),
            audit_deal_sheet(path, "Deals_신규사업,BB", 8),
        ],
    }


def read_xlsx_xml_rows(path: Path) -> dict[str, list[list[str]]]:
    """Read visible cell values directly from xlsx XML.

    The original network workbook has useful sheet XML but openpyxl reports
    empty dimensions for several sheets, so this path avoids relying on
    worksheet dimensions.
    """

    with zipfile.ZipFile(path) as z:
        shared_strings: list[str] = []
        sst_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for item in sst_root.findall("a:si", XML_NS):
            shared_strings.append("".join(item.itertext()))

        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
        output: dict[str, list[list[str]]] = {}

        sheets_node = wb_root.find("a:sheets", XML_NS)
        if sheets_node is None:
            return output

        for sheet in sheets_node:
            sheet_name = sheet.attrib["name"]
            rel_id = sheet.attrib[f"{{{XML_NS['r']}}}id"]
            target = rels[rel_id]
            sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            sheet_root = ET.fromstring(z.read(sheet_path))

            rows: list[list[str]] = []
            for row in sheet_root.findall(".//a:sheetData/a:row", XML_NS):
                values: list[str] = []
                for cell in row.findall("a:c", XML_NS):
                    value_node = cell.find("a:v", XML_NS)
                    value = "".join(value_node.itertext()) if value_node is not None else ""
                    if cell.attrib.get("t") == "s" and value:
                        value = shared_strings[int(value)]
                    if value:
                        values.append(value.strip())
                if values:
                    rows.append(values)
            output[sheet_name] = rows

    return output


def audit_network_original(path: Path) -> dict[str, Any]:
    sheets = read_xlsx_xml_rows(path)
    output = {
        "file": str(path.relative_to(ROOT)),
        "sha256": file_sha256(path),
        "sheets": {},
    }
    for sheet_name, rows in sheets.items():
        joined = "\n".join(" | ".join(row) for row in rows)
        output["sheets"][sheet_name] = {
            "non_empty_rows": len(rows),
            "first_rows": rows[:5],
            "mentions_nda": joined.count("NDA"),
            "mentions_profile": joined.count("프로필"),
            "mentions_appointment": joined.count("위촉"),
            "mentions_account": joined.count("계정"),
        }
    return output


def audit_to_go(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet2"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        text = " | ".join(cell_text(v) for v in row if cell_text(v))
        if text:
            rows.append(text)
    wb.close()

    keywords = ["NDA", "프로필", "출장", "IR", "투자", "M&A", "일정", "계약", "제안서", "입금", "파트너", "행사"]
    return {
        "file": str(path.relative_to(ROOT)),
        "sha256": file_sha256(path),
        "sheet": "Sheet2",
        "non_empty_rows": len(rows),
        "keyword_hits": {keyword: sum(1 for row in rows if keyword.lower() in row.lower()) for keyword in keywords},
        "sample_rows": rows[:30],
    }


def write_markdown(report: dict[str, Any], path: Path) -> None:
    lines = [
        "# XP ERP Source Audit",
        "",
        "Generated by `scripts/audit_sources.py`.",
        "",
        "## Summary",
        "",
        f"- Cleaned partners: {report['cleaned_partners']['rows']} rows, {report['cleaned_partners']['category_counts'].get('(blank)', 0)} blank categories.",
        f"- Original network sheets: {len(report['network_original']['sheets'])} sheets with XML-readable rows.",
        f"- Deal list priority sheets audited: {len(report['deal_list']['sheets'])}.",
        f"- To Go List: {report['to_go']['non_empty_rows']} non-empty operational rows.",
        "",
        "## Cleaned Partner List",
        "",
        "```json",
        json.dumps(report["cleaned_partners"], ensure_ascii=False, indent=2),
        "```",
        "",
        "## Original Network List",
        "",
    ]

    for sheet_name, sheet_report in report["network_original"]["sheets"].items():
        lines.extend(
            [
                f"### {sheet_name}",
                "",
                f"- Non-empty rows: {sheet_report['non_empty_rows']}",
                f"- NDA mentions: {sheet_report['mentions_nda']}",
                f"- Profile mentions: {sheet_report['mentions_profile']}",
                f"- Appointment mentions: {sheet_report['mentions_appointment']}",
                f"- Account mentions: {sheet_report['mentions_account']}",
                "",
            ]
        )

    lines.extend(
        [
            "## Deal List",
            "",
            "```json",
            json.dumps(report["deal_list"], ensure_ascii=False, indent=2),
            "```",
            "",
            "## To Go List",
            "",
            "```json",
            json.dumps(report["to_go"], ensure_ascii=False, indent=2),
            "```",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    missing = [str(path) for path in SOURCES.values() if not path.exists()]
    if missing:
        raise SystemExit("Missing source files:\n" + "\n".join(missing))

    report = {
        "cleaned_partners": audit_cleaned_partners(SOURCES["cleaned_partners"]),
        "deal_list": audit_deal_list(SOURCES["deal_list"]),
        "network_original": audit_network_original(SOURCES["network_original"]),
        "to_go": audit_to_go(SOURCES["to_go"]),
    }

    DOCS.mkdir(exist_ok=True)
    (DOCS / "source-audit.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown(report, DOCS / "source-audit.md")
    print(json.dumps({
        "source_audit_json": str((DOCS / "source-audit.json").relative_to(ROOT)),
        "source_audit_md": str((DOCS / "source-audit.md").relative_to(ROOT)),
        "cleaned_partner_rows": report["cleaned_partners"]["rows"],
        "to_go_rows": report["to_go"]["non_empty_rows"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
