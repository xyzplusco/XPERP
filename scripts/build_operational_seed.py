#!/usr/bin/env python3
"""Build structured operational seed data for the XP ERP app.

This produces a reviewable JSON file used by the current Next.js scaffold.
It is deliberately conservative: uncertain source rows remain visible instead
of being silently normalized away.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW = DATA / "raw"
OUT = DATA / "processed"
DOCS = ROOT / "docs"

XML_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return str(value).strip()


def compact(value: str) -> str:
    return re.sub(r"\s+", "", value.lower())


def status_from_mark(value: str) -> str:
    normalized = value.strip().upper()
    if normalized in {"O", "Y", "YES", "완료", "수령", "SIGNED", "체결"}:
        return "received"
    if normalized in {"X", "N", "NO", "미수령", "미완료"}:
        return "needed"
    if re.search(r"\d{4}[./-]\d{1,2}[./-]\d{1,2}", value):
        return "received"
    if value:
        return "requested"
    return "needed"


def source_ref(source: str, sheet: str, row: int) -> str:
    return f"{source}:{sheet}:{row}"


def load_cleaned_people() -> list[dict[str, str]]:
    path = DATA / "XP_partner_list_cleaned_DB_ready.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Cleaned_Partners"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [text(v) for v in rows[0]]
    idx = {name: i for i, name in enumerate(headers) if name}

    def col(row: tuple[Any, ...], name: str) -> str:
        pos = idx.get(name)
        return text(row[pos]) if pos is not None and pos < len(row) else ""

    people = []
    for source_row, row in enumerate(rows[1:], start=2):
        name = col(row, "name")
        if not name:
            continue
        people.append(
            {
                "id": f"person:{compact(name)}:{source_row}",
                "sourceRef": source_ref("cleaned_partners", "Cleaned_Partners", source_row),
                "name": name,
                "segment": segment_from_category(col(row, "category")),
                "category": col(row, "category") or "Unclassified",
                "company": col(row, "company"),
                "department": col(row, "department"),
                "role": col(row, "title"),
                "email": col(row, "email_personal") or col(row, "email_company"),
                "phone": col(row, "phone_personal") or col(row, "phone_company"),
                "recommender": col(row, "recommender"),
                "owner": col(row, "internal_manager"),
                "ndaStatus": col(row, "nda_status") or "Unknown",
                "profileStatus": "Unknown",
                "appointmentStatus": col(row, "agreement_status") or "Unknown",
                "expertise": "; ".join(
                    part for part in [col(row, "function_experience"), col(row, "industry_experience")] if part
                ),
            }
        )
    wb.close()
    return people


def segment_from_category(category: str) -> str:
    if category in {"임원", "직원"}:
        return "XP internal"
    if category in {"파트너", "파트너 후보", "후보", "파트너 (비활성화)"}:
        return "Partner network"
    if category == "협력사":
        return "Vendor advisor"
    return "Unclassified"


def read_xlsx_xml_rows(path: Path) -> dict[str, list[tuple[int, list[str]]]]:
    with zipfile.ZipFile(path) as z:
        shared_strings: list[str] = []
        sst_root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for item in sst_root.findall("a:si", XML_NS):
            shared_strings.append("".join(item.itertext()))

        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
        sheets_node = wb_root.find("a:sheets", XML_NS)
        if sheets_node is None:
            return {}

        output: dict[str, list[tuple[int, list[str]]]] = {}
        for sheet in sheets_node:
            sheet_name = sheet.attrib["name"]
            rel_id = sheet.attrib[f"{{{XML_NS['r']}}}id"]
            target = rels[rel_id]
            sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            sheet_root = ET.fromstring(z.read(sheet_path))

            rows: list[tuple[int, list[str]]] = []
            for row in sheet_root.findall(".//a:sheetData/a:row", XML_NS):
                row_number = int(row.attrib.get("r", "0") or "0")
                values: list[str] = []
                for cell in row.findall("a:c", XML_NS):
                    value_node = cell.find("a:v", XML_NS)
                    value = "".join(value_node.itertext()) if value_node is not None else ""
                    if cell.attrib.get("t") == "s" and value:
                        value = shared_strings[int(value)]
                    if value:
                        values.append(value.strip())
                if values:
                    rows.append((row_number, values))
            output[sheet_name] = rows
    return output


def load_network_rows() -> list[dict[str, str]]:
    rows_by_sheet = read_xlsx_xml_rows(RAW / "00.XP_파트너 및 네트워크 리스트_260813.xlsx")
    records: list[dict[str, str]] = []

    for sheet, rows in rows_by_sheet.items():
        for row_number, values in rows:
            if values[0] in {"현재 파트너 및 후보군", "As-Is", "기존 DMC"}:
                continue
            if any(header in values[0] for header in ["성명", "이름", "분류", "분야"]):
                continue

            record = parse_network_row(sheet, row_number, values)
            if record:
                records.append(record)

    return records


def parse_network_row(sheet: str, row_number: int, values: list[str]) -> dict[str, str] | None:
    if sheet in {"XP", "컨설팅파트너", "투자_재무 파트너"}:
        if len(values) < 3:
            return None
        name = values[1] if values[0] in {"상", "중", "하"} and len(values) > 1 else values[0]
        if not name or name in {"성명", "권한"}:
            return None
        offset = 1 if values[0] in {"상", "중", "하"} else 0
        return {
            "sourceRef": source_ref("network_original", sheet, row_number),
            "name": name,
            "segment": "XP internal" if sheet == "XP" else "Consulting partner",
            "category": values[offset + 2] if len(values) > offset + 2 else sheet,
            "company": values[offset + 8] if len(values) > offset + 8 else "",
            "role": values[offset + 3] if len(values) > offset + 3 else "",
            "coreField": values[offset + 4] if len(values) > offset + 4 else "",
            "expertise": values[offset + 5] if len(values) > offset + 5 else "",
            "email": values[offset + 6] if len(values) > offset + 6 else "",
            "phone": values[offset + 7] if len(values) > offset + 7 else "",
            "ndaStatus": values[offset + 9] if len(values) > offset + 9 else "",
            "profileStatus": values[offset + 10] if len(values) > offset + 10 else "",
            "appointmentStatus": values[offset + 11] if len(values) > offset + 11 else "",
        }

    if sheet == "LP":
        if len(values) < 2:
            return None
        name = values[1] if len(values) > 1 else values[0]
        return {
            "sourceRef": source_ref("network_original", sheet, row_number),
            "name": name,
            "segment": "LP / investor",
            "category": values[0],
            "company": values[3] if len(values) > 3 else "",
            "role": values[2] if len(values) > 2 else "",
            "coreField": "LP / investor",
            "expertise": values[6] if len(values) > 6 else "",
            "email": values[5] if len(values) > 5 else "",
            "phone": "",
            "ndaStatus": "",
            "profileStatus": "",
            "appointmentStatus": "",
        }

    if sheet == "외부전문가_2603":
        if len(values) < 2:
            return None
        return {
            "sourceRef": source_ref("network_original", sheet, row_number),
            "name": values[0],
            "segment": "External expert",
            "category": "외부전문가",
            "company": values[1] if len(values) > 1 else "",
            "role": values[3] if len(values) > 3 else "",
            "coreField": values[11] if len(values) > 11 else "",
            "expertise": "; ".join(values[11:14]) if len(values) > 11 else "",
            "email": values[4] if len(values) > 4 else "",
            "phone": values[5] if len(values) > 5 else "",
            "ndaStatus": "",
            "profileStatus": "",
            "appointmentStatus": "",
        }

    if sheet == "협력사 리스트":
        if len(values) < 3:
            return None
        return {
            "sourceRef": source_ref("network_original", sheet, row_number),
            "name": values[2],
            "segment": "Vendor advisor",
            "category": values[0],
            "company": values[1],
            "role": values[3] if len(values) > 3 else "",
            "coreField": values[0],
            "expertise": values[0],
            "email": values[4] if len(values) > 4 else "",
            "phone": values[5] if len(values) > 5 else "",
            "ndaStatus": values[6] if len(values) > 6 else "",
            "profileStatus": "",
            "appointmentStatus": values[7] if len(values) > 7 else "",
        }

    return None


def merge_network(cleaned: list[dict[str, str]], original: list[dict[str, str]]) -> list[dict[str, str]]:
    merged: dict[str, dict[str, str]] = {}
    for record in cleaned:
        key = compact(record["name"])
        merged[key] = record

    for record in original:
        key = compact(record["name"])
        if not key:
            continue
        if key in merged:
            current = merged[key]
            for field in [
                "segment",
                "category",
                "company",
                "role",
                "email",
                "phone",
                "ndaStatus",
                "profileStatus",
                "appointmentStatus",
                "expertise",
            ]:
                if record.get(field) and (not current.get(field) or current.get(field) == "Unknown"):
                    current[field] = record[field]
            current["sourceRef"] = f"{current['sourceRef']}; {record['sourceRef']}"
        else:
            merged[key] = {
                "id": f"person:{key}",
                "sourceRef": record["sourceRef"],
                "name": record["name"],
                "segment": record.get("segment", "Unclassified"),
                "category": record.get("category", ""),
                "company": record.get("company", ""),
                "department": "",
                "role": record.get("role", ""),
                "email": record.get("email", ""),
                "phone": record.get("phone", ""),
                "recommender": "",
                "owner": "",
                "ndaStatus": record.get("ndaStatus", "") or "Unknown",
                "profileStatus": record.get("profileStatus", "") or "Unknown",
                "appointmentStatus": record.get("appointmentStatus", "") or "Unknown",
                "expertise": record.get("expertise", ""),
            }
    return list(merged.values())


def load_projects() -> list[dict[str, str]]:
    path = DATA / "XP Deal list_대외비_20260805.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    projects: list[dict[str, str]] = []
    sheets = [("Deals_0731", 7), ("M&A Deal_0809", 3), ("Deals_신규사업,BB", 8), ("Deals_투자매각", 5)]

    for sheet_name, header_row in sheets:
        ws = wb[sheet_name]
        headers = [text(c.value) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
        idx = {name: i for i, name in enumerate(headers) if name}

        def col(row: tuple[Any, ...], name: str) -> str:
            pos = idx.get(name)
            return text(row[pos]) if pos is not None and pos < len(row) else ""

        update_cols = [(name, pos) for name, pos in idx.items() if re.search(r"\d+월|주", name)]
        for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            company = col(row, "회사명")
            if not company:
                continue
            updates = [text(row[pos]) for _, pos in update_cols if pos < len(row) and text(row[pos])]
            latest_update = updates[-1] if updates else ""
            projects.append(
                {
                    "sourceRef": source_ref("deal_list", sheet_name, source_row),
                    "company": company,
                    "representative": col(row, "대표자"),
                    "projectType": map_project_type(col(row, "서비스 섹터") or sheet_name),
                    "sourceType": col(row, "서비스 섹터") or sheet_name,
                    "business": col(row, "사업"),
                    "pl": col(row, "PL"),
                    "pm": col(row, "PM 1") or col(row, "PM2"),
                    "clientNeed": col(row, "대표 니즈"),
                    "xpRequest": col(row, "XP 요청"),
                    "contractStatus": col(row, "계약 현황") or col(row, "매각여부") or "Review",
                    "latestUpdate": latest_update[:300],
                    "weeklyUpdateCount": str(len(updates)),
                    "nextAction": infer_project_next_action(company, latest_update, col(row, "XP 요청")),
                }
            )
    wb.close()
    return projects


def map_project_type(value: str) -> str:
    normalized = value.lower()
    if any(token in normalized for token in ["bpr", "리엔지니어링", "ax"]):
        return "Re-engineering"
    if any(token in normalized for token in ["bb", "비즈니스"]):
        return "Business building"
    if any(token in normalized for token in ["fim", "투자", "매각", "m&a"]):
        return "Investment / M&A"
    if any(token in normalized for token in ["gx", "해외"]):
        return "Go Global"
    if "영업" in normalized or "컨설팅" in normalized:
        return "Consulting"
    return "Review"


def infer_project_next_action(company: str, latest_update: str, xp_request: str) -> str:
    if latest_update:
        return latest_update[:120]
    if xp_request:
        return xp_request[:120]
    return f"{company} source row review"


def load_to_go_tasks() -> list[dict[str, str]]:
    path = RAW / "To Go List XYZ Plus (7).xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet2"]
    tasks: list[dict[str, str]] = []
    action_words = ("일정", "계약", "제안서", "NDA", "프로필", "확정", "미팅", "공유", "인수인계", "입금", "출장", "IR")

    for source_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [text(v) for v in row if text(v)]
        if not cells:
            continue
        raw = " | ".join(cells)
        classification = "likely_task" if any(word.lower() in raw.lower() for word in action_words) else "needs_review"
        if len(cells) == 1 and len(cells[0]) < 12 and source_row < 10:
            classification = "heading"
        tasks.append(
            {
                "sourceRef": source_ref("to_go", "Sheet2", source_row),
                "title": cells[0],
                "body": raw,
                "owner": infer_owner(cells),
                "linkedArea": infer_linked_area(raw),
                "status": "Review" if classification == "needs_review" else "Backlog",
                "classification": classification,
            }
        )
    wb.close()
    return tasks


def infer_owner(cells: list[str]) -> str:
    for value in reversed(cells):
        if re.fullmatch(r"[가-힣]{2,4}", value):
            return value
    return cells[-1] if len(cells) > 1 else "Unassigned"


def infer_linked_area(raw: str) -> str:
    if any(token in raw for token in ["NDA", "프로필", "계약", "MOU"]):
        return "Documents"
    if any(token in raw for token in ["IR", "출장", "행사", "일정"]):
        return "Events"
    if any(token in raw for token in ["투자", "M&A", "제안서", "입금"]):
        return "Projects"
    return "Network"


def build_document_requirements(network: list[dict[str, str]], tasks: list[dict[str, str]]) -> list[dict[str, str]]:
    requirements: list[dict[str, str]] = []

    for person in network:
        if person.get("segment") in {"Unclassified", "External expert"}:
            continue
        for field, req_type in [
            ("ndaStatus", "NDA"),
            ("profileStatus", "Profile"),
            ("appointmentStatus", "Partner appointment"),
        ]:
            raw_status = person.get(field, "")
            status = status_from_mark(raw_status)
            if status == "needed" or raw_status in {"Unknown", ""}:
                requirements.append(
                    {
                        "subject": person["name"],
                        "type": req_type,
                        "owner": person.get("owner") or "Operations",
                        "status": "Needed" if status == "needed" else "Review",
                        "due": "Review",
                        "sourceRef": person.get("sourceRef", ""),
                    }
                )

    for task in tasks:
        body = task["body"]
        for req_type in ["NDA", "프로필", "계약", "MOU"]:
            if req_type.lower() in body.lower():
                requirements.append(
                    {
                        "subject": task["title"],
                        "type": "Profile" if req_type == "프로필" else req_type,
                        "owner": task["owner"],
                        "status": "Needed",
                        "due": "From To Go",
                        "sourceRef": task["sourceRef"],
                    }
                )

    return requirements


def main() -> None:
    cleaned = load_cleaned_people()
    original = load_network_rows()
    network = merge_network(cleaned, original)
    projects = load_projects()
    tasks = load_to_go_tasks()
    document_requirements = build_document_requirements(network, tasks)

    summary = {
        "people": len(network),
        "projects": len(projects),
        "tasks": len(tasks),
        "documentRequirements": len(document_requirements),
        "sources": 4,
        "networkSegments": dict(Counter(row["segment"] for row in network).most_common()),
        "projectTypes": dict(Counter(row["projectType"] for row in projects).most_common()),
    }

    seed = {
        "summary": summary,
        "network": network,
        "projects": projects,
        "tasks": tasks,
        "documentRequirements": document_requirements,
    }
    OUT.mkdir(parents=True, exist_ok=True)
    output_path = OUT / "operational_seed_preview.json"
    output_path.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_summary(seed, DOCS / "operational-seed.md")
    print(json.dumps({"output": str(output_path.relative_to(ROOT)), **summary}, ensure_ascii=False))


def write_markdown_summary(seed: dict[str, Any], path: Path) -> None:
    summary = seed["summary"]
    lines = [
        "# XP ERP Operational Seed Preview",
        "",
        "Generated by `scripts/build_operational_seed.py`.",
        "",
        "## Counts",
        "",
        f"- People: {summary['people']}",
        f"- Projects: {summary['projects']}",
        f"- Tasks: {summary['tasks']}",
        f"- Document requirements: {summary['documentRequirements']}",
        f"- Source workbooks: {summary['sources']}",
        "",
        "## Network Segments",
        "",
    ]
    for segment, count in summary["networkSegments"].items():
        lines.append(f"- {segment}: {count}")

    lines.extend(["", "## Project Types", ""])
    for project_type, count in summary["projectTypes"].items():
        lines.append(f"- {project_type}: {count}")

    lines.extend(
        [
            "",
            "## Review Notes",
            "",
            "- `Unclassified` people are intentionally preserved for later cleanup.",
            "- `Review` project types mean the source sector needs manual mapping.",
            "- Document requirements are generated from network onboarding fields and To Go List mentions.",
            "- This is a seed preview, not a final production import.",
            "",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
