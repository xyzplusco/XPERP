#!/usr/bin/env python3
"""Create reviewable seed previews from XP ERP source workbooks.

This script intentionally writes CSV previews instead of inserting into a DB.
The previews make parsing assumptions visible before the real importer is wired
to Supabase/Postgres.
"""

from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW = DATA / "raw"
OUT = DATA / "processed"


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return str(value).strip()


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", "", value.lower())


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def cleaned_partner_people() -> list[dict[str, str]]:
    path = DATA / "XP_partner_list_cleaned_DB_ready.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Cleaned_Partners"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [text(v) for v in rows[0]]
    idx = {name: i for i, name in enumerate(headers) if name}

    def col(row: tuple[Any, ...], name: str) -> str:
        pos = idx.get(name)
        return text(row[pos]) if pos is not None and pos < len(row) else ""

    output = []
    for source_row, row in enumerate(rows[1:], start=2):
        name = col(row, "name")
        if not name:
            continue
        output.append(
            {
                "source": "cleaned_partners",
                "source_sheet": "Cleaned_Partners",
                "source_row": str(source_row),
                "name": name,
                "normalized_name": normalize_name(name),
                "category": col(row, "category"),
                "company": col(row, "company"),
                "department": col(row, "department"),
                "title": col(row, "title"),
                "email": col(row, "email_personal") or col(row, "email_company"),
                "phone": col(row, "phone_personal") or col(row, "phone_company"),
                "recommender": col(row, "recommender"),
                "nda_status": col(row, "nda_status"),
                "agreement_status": col(row, "agreement_status"),
                "agreement_end_date": col(row, "agreement_end_date"),
                "internal_manager": col(row, "internal_manager"),
                "expertise_functions": col(row, "function_experience"),
                "expertise_industries": col(row, "industry_experience"),
            }
        )
    wb.close()
    return output


def deal_projects() -> list[dict[str, str]]:
    path = DATA / "XP Deal list_대외비_20260805.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output: list[dict[str, str]] = []
    sheets = [("Deals_0731", 7), ("M&A Deal_0809", 3), ("Deals_신규사업,BB", 8)]

    for sheet_name, header_row in sheets:
        ws = wb[sheet_name]
        headers = [text(c.value) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
        idx = {name: i for i, name in enumerate(headers) if name}

        def col(row: tuple[Any, ...], name: str) -> str:
            pos = idx.get(name)
            return text(row[pos]) if pos is not None and pos < len(row) else ""

        update_cols = [(name, pos) for name, pos in idx.items() if re.search(r"\d+월|주", name)]
        for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            company = col(row, "회사명")
            if not company:
                continue
            updates = []
            for label, pos in update_cols:
                value = text(row[pos]) if pos < len(row) else ""
                if value:
                    updates.append(f"{label}: {value}")
            output.append(
                {
                    "source": "deal_list",
                    "source_sheet": sheet_name,
                    "source_row": str(source_row),
                    "company": company,
                    "representative": col(row, "대표자"),
                    "service_sector": col(row, "서비스 섹터"),
                    "business": col(row, "사업"),
                    "pl": col(row, "PL"),
                    "pm1": col(row, "PM 1"),
                    "pm2": col(row, "PM2"),
                    "client_need": col(row, "대표 니즈"),
                    "xp_request": col(row, "XP 요청"),
                    "contract_status": col(row, "계약 현황") or col(row, "매각여부"),
                    "weekly_updates_joined": "\n".join(updates[:12]),
                }
            )
    wb.close()
    return output


def to_go_tasks() -> list[dict[str, str]]:
    path = RAW / "To Go List XYZ Plus (7).xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet2"]
    output: list[dict[str, str]] = []
    action_words = ("일정", "계약", "제안서", "NDA", "프로필", "확정", "미팅", "공유", "인수인계", "입금", "출장", "IR")

    for source_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [text(v) for v in row if text(v)]
        if not cells:
            continue
        raw = " | ".join(cells)
        if len(cells) == 1 and len(cells[0]) < 12 and source_row < 10:
            status = "heading"
        elif any(word.lower() in raw.lower() for word in action_words):
            status = "likely_task"
        else:
            status = "needs_review"
        output.append(
            {
                "source": "to_go",
                "source_sheet": "Sheet2",
                "source_row": str(source_row),
                "title_guess": cells[0],
                "body": raw,
                "owner_hint": cells[-1] if len(cells) >= 2 else "",
                "classification": status,
            }
        )
    wb.close()
    return output


def main() -> None:
    people = cleaned_partner_people()
    projects = deal_projects()
    tasks = to_go_tasks()

    write_csv(
        OUT / "people_preview.csv",
        people,
        [
            "source",
            "source_sheet",
            "source_row",
            "name",
            "normalized_name",
            "category",
            "company",
            "department",
            "title",
            "email",
            "phone",
            "recommender",
            "nda_status",
            "agreement_status",
            "agreement_end_date",
            "internal_manager",
            "expertise_functions",
            "expertise_industries",
        ],
    )
    write_csv(
        OUT / "project_preview.csv",
        projects,
        [
            "source",
            "source_sheet",
            "source_row",
            "company",
            "representative",
            "service_sector",
            "business",
            "pl",
            "pm1",
            "pm2",
            "client_need",
            "xp_request",
            "contract_status",
            "weekly_updates_joined",
        ],
    )
    write_csv(
        OUT / "task_preview.csv",
        tasks,
        ["source", "source_sheet", "source_row", "title_guess", "body", "owner_hint", "classification"],
    )

    print(
        {
            "people_preview": len(people),
            "project_preview": len(projects),
            "task_preview": len(tasks),
            "output_dir": str(OUT.relative_to(ROOT)),
        }
    )


if __name__ == "__main__":
    main()

